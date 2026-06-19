#!/usr/bin/env python3

from __future__ import annotations

import csv
import re
from pathlib import Path

import openpyxl


ROOT = Path.cwd() / "TMEM158_CaUPR_ESCC"
SOURCE = ROOT / "02_data" / "external" / "spatial_progression" / "natcomm2023_escc_spatial_source_data.xlsx"
OUT = ROOT / "04_results" / "spatial_progression"


def stage_from_label(label: object) -> str:
    text = str(label or "").strip()
    text = re.sub(r"[-_]\d+$", "", text)
    first = text.split()[0] if text else ""
    compact = re.match(r"^(Normal|Nomal|NE|LGIN|HGIN|ESCC)", first, flags=re.IGNORECASE)
    if compact:
        first = compact.group(1)
    mapping = {
        "Normal": "NE",
        "Nomal": "NE",
        "NE": "NE",
        "LGIN": "LGIN",
        "HGIN": "HGIN",
        "ESCC": "ESCC",
    }
    return mapping.get(first, first)


def as_float(value: object):
    try:
        if value is None or value == "":
            return None
        return float(value)
    except Exception:
        return None


def write_rows(path: Path, fieldnames: list[str], rows: list[dict]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def extract_figure_2d(wb) -> list[dict]:
    ws = wb["Figure 2d"]
    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        label = row[0]
        if label is None:
            continue
        stage = stage_from_label(label)
        if stage not in {"NE", "LGIN", "HGIN", "ESCC"}:
            continue
        rows.append(
            {
                "source_sheet": "Figure 2d",
                "sample_label": str(label),
                "stage": stage,
                "macrophages": as_float(row[1]),
                "fibroblasts": as_float(row[2]),
            }
        )
    return rows


def extract_figure_2e(wb) -> list[dict]:
    ws = wb["Figure 2e"]
    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        label = row[0]
        if label is None:
            continue
        stage = stage_from_label(label)
        if stage not in {"NE", "LGIN", "HGIN", "ESCC"}:
            continue
        rows.append(
            {
                "source_sheet": "Figure 2e",
                "sample_label": str(label),
                "stage": stage,
                "macrophages_m0": as_float(row[1]),
                "macrophages_m1": as_float(row[2]),
                "macrophages_m2": as_float(row[3]),
            }
        )
    return rows


def extract_figure_2f(wb) -> list[dict]:
    ws = wb["Figure 2f"]
    rows: list[dict] = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        sample = row[0]
        if sample is not None:
            for idx, stage in enumerate(["NE", "LGIN", "HGIN", "ESCC"], start=1):
                value = as_float(row[idx])
                if value is not None:
                    rows.append(
                        {
                            "source_sheet": "Figure 2f",
                            "sample_label": str(sample),
                            "stage": stage,
                            "marker": "alpha_SMA_fibroblast_IF",
                            "value": value,
                        }
                    )
        sample2 = row[6]
        if sample2 is not None:
            for idx, stage in enumerate(["NE", "LGIN", "HGIN", "ESCC"], start=7):
                value = as_float(row[idx])
                if value is not None:
                    rows.append(
                        {
                            "source_sheet": "Figure 2f",
                            "sample_label": str(sample2),
                            "stage": stage,
                            "marker": "CD68_macrophage_IF",
                            "value": value,
                        }
                    )
    return rows


def extract_roi_markers(wb) -> list[dict]:
    ws = wb["Fig5c, Supplment Fig 5a,9a, b"]
    header = [str(x) if x is not None else "" for x in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        item = {header[i]: row[i] if i < len(row) else None for i in range(len(header))}
        stage = stage_from_label(item.get("his"))
        if stage not in {"NE", "LGIN", "HGIN", "ESCC"}:
            continue
        base = {
            "source_sheet": "Fig5c, Supplment Fig 5a,9a, b",
            "scan_label": str(item.get("ScanLabel")),
            "roi_label": str(item.get("ROILabel")),
            "stage": stage,
        }
        for gene in ["KRT16", "KRT17", "MAL", "TAGLN2", "CRNN"]:
            rows.append({**base, "gene": gene, "expression": as_float(item.get(gene))})
    return rows


def main() -> int:
    OUT.mkdir(parents=True, exist_ok=True)
    if not SOURCE.exists():
        raise FileNotFoundError(f"Missing source workbook: {SOURCE}")
    wb = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)

    cell_abundance = extract_figure_2d(wb)
    macrophage_subtypes = extract_figure_2e(wb)
    if_markers = extract_figure_2f(wb)
    roi_markers = extract_roi_markers(wb)

    write_rows(
        OUT / "natcomm2023_spatial_cell_abundance.csv",
        ["source_sheet", "sample_label", "stage", "macrophages", "fibroblasts"],
        cell_abundance,
    )
    write_rows(
        OUT / "natcomm2023_spatial_macrophage_subtypes.csv",
        ["source_sheet", "sample_label", "stage", "macrophages_m0", "macrophages_m1", "macrophages_m2"],
        macrophage_subtypes,
    )
    write_rows(
        OUT / "natcomm2023_spatial_if_markers.csv",
        ["source_sheet", "sample_label", "stage", "marker", "value"],
        if_markers,
    )
    write_rows(
        OUT / "natcomm2023_spatial_roi_marker_expression.csv",
        ["source_sheet", "scan_label", "roi_label", "stage", "gene", "expression"],
        roi_markers,
    )
    write_rows(
        OUT / "natcomm2023_spatial_source_extract_status.csv",
        ["field", "value"],
        [
            {"field": "module_status", "value": "completed"},
            {"field": "source_workbook", "value": str(SOURCE)},
            {"field": "source_article", "value": "Liu et al. Nature Communications 2023, doi:10.1038/s41467-023-40343-5"},
            {"field": "cell_abundance_rows", "value": len(cell_abundance)},
            {"field": "macrophage_subtype_rows", "value": len(macrophage_subtypes)},
            {"field": "if_marker_rows", "value": len(if_markers)},
            {"field": "roi_marker_rows", "value": len(roi_markers)},
            {"field": "full_wta_matrix_available_in_source_data", "value": "FALSE"},
            {"field": "direct_TAC_high_score_possible", "value": "FALSE"},
        ],
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
